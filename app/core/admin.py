"""
Django admin customization.
"""
from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.utils.translation import gettext_lazy as _

from .models import Ride, RideEvent

from .utils import get_long_trips_report
from django.urls import path, reverse
from django.http import HttpResponse
from django.shortcuts import render
from core import models

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

class UserAdmin(BaseUserAdmin):
    """Define the admin pages for users."""

    ordering = ['id_user']
    list_display = ['email', 'first_name', 'last_name', 'role']
    list_filter = ['role', 'is_active', 'is_staff']
    
    fieldsets = (
        (None, {'fields': ('email', 'password')}),
        (
            _('Personal Info'),
            {
                'fields': (
                    'first_name',
                    'last_name', 
                    'phone_number',
                    'role',
                )
            }
        ),
        (
            _('Permissions'),
            {
                'fields': (
                    'is_active',
                    'is_staff',
                    'is_superuser',
                )
            }
        ),
        (_('Important dates'), {'fields': ('last_login',)}),
    )
    readonly_fields = ['last_login']
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': (
                'email',
                'password1',
                'password2',
                'first_name',
                'last_name',
                'phone_number',
                'role',
                'is_active',
                'is_staff',
                'is_superuser',
            )
        }),
    )
class RideEventInline(admin.TabularInline):
    model = RideEvent
    extra = 0
    readonly_fields = ('description', 'created_at')
    can_delete = False

@admin.register(Ride)
class RideAdmin(admin.ModelAdmin):
    list_display = ('id_ride', 'status', 'id_rider', 'id_driver', 'pickup_time')
    list_filter = ('status',)
    inlines = [RideEventInline]
    
    change_list_template = 'admin/ride/ride_changelist.html'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('download-long-trips-report/', 
                 self.admin_site.admin_view(self.download_excel_report),
                 name='ride_long_trips_report'),
        ]
        return custom_urls + urls
    
    def download_excel_report(self, request):
        """Download the long trips report as Excel file."""
        # Get report data
        report_data = get_long_trips_report()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Long Trips Report"
        
        # Add title
        ws.merge_cells('A1:C1')
        ws['A1'] = 'Long Trips Report (> 1 hour)'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Add headers
        headers = ['Month', 'Driver', 'Count of Trips > 1 hr']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data
        for row_idx, row in enumerate(report_data, 4):
            ws.cell(row=row_idx, column=1, value=row['month'])
            ws.cell(row=row_idx, column=2, value=row['driver_name'])
            ws.cell(row=row_idx, column=3, value=row['trip_count'])
            
            # Add borders to data cells
            for col in range(1, 4):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                if col == 3:  # Center align count column
                    cell.alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        
        # Save to BytesIO object
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="long_trips_report.xlsx"'
        
        return response

@admin.register(RideEvent)
class RideEventAdmin(admin.ModelAdmin):
    list_display = ('id_ride_event', 'id_ride', 'description', 'created_at')
    list_filter = ('created_at', 'id_ride__status')
    search_fields = ('description', 'id_ride__id_ride')
    date_hierarchy = 'created_at'
    
    # Make all fields editable including created_at
    fields = ('id_ride', 'description', 'created_at')
    
    # Better formatting for the list
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('id_ride', 'id_ride__id_driver', 'id_ride__id_rider')
    
    # Custom admin actions
    actions = ['set_pickup_time', 'set_dropoff_time']
    
    def set_pickup_time(self, request, queryset):
        """Quick action to set events as pickup events."""
        for event in queryset:
            event.description = "Status changed from 'en-route' to 'pickup'"
            event.save()
        self.message_user(request, f"Updated {queryset.count()} events to pickup")
    
    def set_dropoff_time(self, request, queryset):
        """Quick action to set events as dropoff events."""
        for event in queryset:
            event.description = "Status changed from 'pickup' to 'dropoff'"
            event.save()
        self.message_user(request, f"Updated {queryset.count()} events to dropoff")
    
    set_pickup_time.short_description = "Set as pickup event"
    set_dropoff_time.short_description = "Set as dropoff event"

admin.site.register(models.User, UserAdmin)

